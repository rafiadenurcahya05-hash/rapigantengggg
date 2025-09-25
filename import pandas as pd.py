import pandas as pd
import json
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def convert_json_to_excel_with_multiple_sheets(json_file_path, excel_output_path):
    """
    Membaca file JSON dengan struktur bersarang, mengonversi setiap sub-kategori
    menjadi sheet terpisah dalam satu file Excel, dan merapikannya.
    """
    print(f"Debug: Mencoba membaca file: {json_file_path}")
    print(f"Debug: Direktori saat ini: {os.getcwd()}")
    
    if not os.path.exists(json_file_path):
        print(f"Error: File '{json_file_path}' tidak ditemukan. Pastikan file ada di folder yang sama.")
        return
    
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        print("Debug: JSON berhasil dimuat.")
    except json.JSONDecodeError as e:
        print(f"Error: Gagal membaca JSON - {e}. Periksa format file.")
        return
    except Exception as e:
        print(f"Error tak terduga saat membaca file: {e}")
        return

    if "demo_indonesia" not in data:
        print("Error: Kunci 'demo_indonesia' tidak ditemukan dalam file JSON.")
        return

    demo_data = data["demo_indonesia"]
    print(f"Debug: Ditemukan {len(demo_data)} kategori: {list(demo_data.keys())}")

    try:
        with pd.ExcelWriter(excel_output_path, engine='openpyxl') as writer:
            for category_name, category_list in demo_data.items():
                print(f"Debug: Memproses kategori '{category_name}'...")
                
                if isinstance(category_list, list) and category_list:
                    df = pd.DataFrame(category_list)
                    df_rapi = df.sort_values(by='id').reset_index(drop=True)
                    print(f"Debug: DataFrame untuk '{category_name}' memiliki {len(df_rapi)} baris.")

                    sheet_name = category_name.replace('_', ' ').title()
                    # Batasi nama sheet agar <31 karakter
                    if len(sheet_name) > 31:
                        sheet_name = sheet_name[:28] + '...'
                    
                    df_rapi.to_excel(writer, sheet_name=sheet_name, index=False)

                    # Akses worksheet dengan try-except untuk handle error formatting
                    try:
                        workbook = writer.book
                        worksheet = writer.sheets[sheet_name]

                        header_font = Font(bold=True)
                        for col_idx, column_name in enumerate(df_rapi.columns, 1):
                            # Header bold
                            cell = worksheet.cell(row=1, column=col_idx)
                            cell.font = header_font
                            
                            # Auto-width
                            max_length = max(len(str(column_name)), 
                                           max([len(str(val)) for val in df_rapi[column_name]] + [0]))
                            adjusted_width = min(max_length + 2, 50)  # Batasi max width 50
                            worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
                        
                        print(f"Debug: Formatting selesai untuk '{sheet_name}'.")
                    except Exception as format_error:
                        print(f"Peringatan: Error formatting sheet '{sheet_name}': {format_error}. Data tetap tersimpan.")
                else:
                    print(f"Peringatan: Kategori '{category_name}' kosong atau bukan list. Dilewati.")
    except Exception as excel_error:
        print(f"Error saat membuat Excel: {excel_error}")
        return

    print(f"\nKonversi berhasil! File Excel '{excel_output_path}' telah dibuat dengan data di sheet terpisah.")

# --- Penggunaan ---
if __name__ == "__main__":
    # Gunakan raw string untuk handle spasi di nama file
    json_input_file = r'RAFI ADE NURCAHYA_V3925014.json'  # Ganti dengan path lengkap jika perlu
    excel_output_file = 'Data_Demo_Indonesia.xlsx'

    convert_json_to_excel_with_multiple_sheets(json_input_file, excel_output_file)

    # Preview tanpa bergantung pada file (gunakan data dari fungsi jika perlu)
    print("\n--- Preview Data Kominfo ---")
    try:
        with open(json_input_file, 'r', encoding='utf-8') as f:
            data_preview = json.load(f)
        if "demo_indonesia" in data_preview and "kominfo" in data_preview["demo_indonesia"]:
            df_kominfo_preview = pd.DataFrame(data_preview["demo_indonesia"]["kominfo"])
            print(df_kominfo_preview.head())
        else:
            print("Data Kominfo tidak ditemukan untuk preview.")
    except Exception as e:
        print(f"Gagal menampilkan preview: {e}")
